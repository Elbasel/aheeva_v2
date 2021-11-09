import openpyxl
import pandas as pd
import csv
import requests
import os
from datetime import datetime
from shutil import copy
import calendar


def removeFormatting(ws):
    # ws is not the worksheet name, but the worksheet object
    for row in ws.iter_rows():
        for cell in row:
            cell.style = 'Normal'


def get_creds():
    creds = {}
    with open('credentials.csv') as creds_csv:
        csv_reader = csv.reader(creds_csv)
        for row in csv_reader:
            creds[row[0].lower()] = {'username': row[1], 'password': row[2]}
    return creds


def update_data(account, start_date, end_date):
    print(f'Updating data: {account} {start_date} - {end_date}')

    try:

        copy(f'data_cache/{account}_{start_date}_{end_date}_inbound.csv', f'{account}_inbound.csv')
        copy(f'data_cache/{account}_{start_date}_{end_date}_outbound.csv', f'{account}_outbound.csv')
        print('data cache found')
        print('loading data from cache')
        return
    except FileNotFoundError:
        pass

    try:
        credentials = get_creds()
    except FileNotFoundError as ex:
        print(ex)
        return

    # Getting file from aheeva
    try:
        login_data = {
            "username": credentials[account]['username'],
            "password": credentials[account]['password'],
            "realm": account
        }
    except KeyError as ex:
        print(f'\nCredentials for {account.capitalize()} was not found in credentials file.')
        return 1

    account_id = {
        'furless': 8,
        'bat': 4,
        'rizkalla': 7,
        'swvl': 9,
        'udacity': 6,
        'zoomcar': 11
    }

    today_date = datetime.today()

    day = today_date.day

    start_day = start_date.day
    start_day += 1

    end_day = end_date.day
    end_month = str(end_date.month).zfill(2)
    end_year = str(end_date.year)

    time_start = '04'
    time_end = '03'
    offset = '-4'

    agent_ids = ''
    start_month = start_date.month
    start_year = start_date.year
    if account == 'zoomcar':
        start_day -= 2

        # print()
        # print(f'{start_day=}')
        # print()
        # start_day = str(start_day).zfill(2)
        time_start = '22'
        time_end = '21'
        offset = '2'
        end_day -= 1
        # agent_ids = '12953,13415,13530,13531,13533,13534,13536,13537,13538,13539,13540,13541,13542,13543,13544,13545,13546,13547,13550,13551,13552,13553,13554,13555,13556,13557,13558,13560,13561,13562,13594,13595,13596,13806,13807,13808,66611,66622,66633,66644,8888,99991,99992'

    if account == 'udacity':
        time_start = '05'
        time_end = '04'
        offset = '-5'

    if start_day == 0:
        start_month -= 1
        # print()
        # print(f'{start_month=}')
        # print()
        start_day = calendar.monthrange(start_year, start_month)[1]

    start_month = str(start_month).zfill(2)
    # start_year = str(start_year)

    inbound_urls = {
        'furless':
            f'https://centrogs.sencall-technologies.com:8443/aheevaManagerServer/api/getInboundCallLifeDownload?tenantDbid=8&fromDate={start_year}-{start_month}-{str(start_day - 1).zfill(2)}T{time_start}:00:00.000Z&toDate={end_year}-{end_month}-{str(end_day + 1).zfill(2)}T{time_end}:59:59.000Z&offset={offset}&tenantName=Furless&queueList=CSAT_TEST,Inquiries,Inquiries_Q,In_Convolo,In_FurlessCSAT_ARABIC,In_FurlessCSAT_ENGLISH,In_Furless_Callback,In_Furless_Voice,Fur_1EN_1_AppBook_Q,Fur_1EN_2_Electrolysis_Q,Fur_1EN_3_Feedback_Q,Fur_2AR_1_AppBook_Q,Fur_2AR_2_Electrolysis_Q,Fur_2AR_3_Feedback_Q,Support,Support_Q&language=EN&rollBack=&type=EXCEL&schedule=',
        'rizkalla':
            f'https://centrogs.sencall-technologies.com:8443/aheevaManagerServer/api/getInboundCallLifeDownload?tenantDbid=7&fromDate={start_year}-{start_month}-{str(start_day - 1).zfill(2)}T{time_start}:00:00.000Z&toDate={end_year}-{end_month}-{str(end_day + 1).zfill(2)}T{time_end}:59:59.000Z&offset={offset}&tenantName=Rizkalla&queueList=In_Rizkalla,Rizkalla_1_Inquiries_Q,Rizkalla_2_Installments_Q,Rizkalla_3_Complaints_Q,IN_RizkCSAT&language=EN&rollBack=&type=EXCEL&schedule=',
        'bat':
            f'https://centrogs.sencall-technologies.com:8443/aheevaManagerServer/api/getInboundCallLifeDownload?tenantDbid=4&fromDate={start_year}-{start_month}-{str(start_day - 1).zfill(2)}T{time_start}:00:00.000Z&toDate={end_year}-{end_month}-{str(end_day + 1).zfill(2)}T{time_end}:59:59.000Z&offset={offset}&tenantName=BAT&queueList=Inbound_BAT,BAT_1AR_1CS_Queue,BAT_1AR_2QA_Queue,BAT_1AR_3Sales_Queue,BAT_2EN_1CS_Queue,BAT_2EN_2QA_Queue,BAT_2EN_3Sales_Queue,BAT_1AR_1CS_VM,BAT_1AR_2QA_VM,BAT_1AR_3Sales_VM,BAT_2EN_1CS_VM,BAT_2EN_2QA_VM,BAT_2EN_3Sales_VM&language=EN&rollBack=&type=EXCEL&schedule=',
        'swvl':
            f'https://centrogs.sencall-technologies.com:8443/aheevaManagerServer/api/getInboundCallLifeDownload?tenantDbid=9&fromDate={start_year}-{start_month}-{str(start_day - 1).zfill(2)}T{time_start}:00:00.000Z&toDate={end_year}-{end_month}-{str(end_day + 1).zfill(2)}T{time_end}:59:59.000Z&offset={offset}&tenantName=SWVL&queueList=Broadcast_Scheduled,Detected_Machine_Q,Detected_Human_Q,Inbound_Swvl_Help_Line,Swvl_Help_Line_Q,Inbound_Swvl_v.2,Rider_1AR_4LateCancel_Q,Out_of_Hours_Q,Rider_1AR_1Gen_Q,Rider_1AR_2CapIssue_Q,Rider_1AR_3Concerns_Q,Rider_2EN_1Gen_Q,Rider_2EN_2CapIssue_Q,Rider_2EN_3Concerns_Q,Rider_2EN_4LateCancel_Q,Admin_AR_Queue,Admin_EN_Queue,IN_SWVL_CS_CSAT&language=EN&rollBack=&type=EXCEL&schedule=',
        'udacity':
            f'https://centrogs.sencall-technologies.com:8443/aheevaManagerServer/api/getInboundCallLifeDownload?tenantDbid=6&fromDate={start_year}-{start_month}-{str(start_day - 1).zfill(2)}T{time_start}:00:00.000Z&toDate={end_year}-{end_month}-{str(end_day + 1).zfill(2)}T{time_end}:59:59.000Z&offset={offset}&tenantName=Udacity&queueList=In_CSATSurvey,In_Udacity&language=EN&rollBack=&type=EXCEL&schedule=',
        'zoomcar':
            f'https://centrogs.sencall-technologies.com:8443/aheevaManagerServer/api/getInboundCallLifeDownload?tenantDbid=11&fromDate={start_year}-{start_month}-{str(start_day).zfill(2)}T{time_start}:00:00.000Z&toDate={end_year}-{end_month}-{str(end_day + 1).zfill(2)}T{time_end}:59:59.000Z&offset={offset}&tenantName=ZoomCar&queueList=In_Zoomcar,OOH_Zoomcar_Q,AR_1Emerg_Q,AR_2Rental_3Locate_Q,AR_2Rental_1Info_5MM1_Q,AR_2Rental_2Price_5MM1_Q,AR_2Rental_1Info_0CSR_Q,AR_2Rental_2Price_0CSR_Q,AR_2Rental_4Paper_Q,AR_2Rental_0CSR_Q,EN_1Emerg_Q,EN_2Rental_1Info_0CSR_Q,EN_2Rental_1Info_5MM1_Q,EN_2Rental_2Price_5MM1_Q,EN_2Rental_2Price_0CSR_Q,EN_2Rental_3Locate_Q,EN_2Rental_4Paper,EN_2Rental_0CSR_Q,EN_3Owner_Q,AR_3Owner_Q,Zoomcar_AR_CSAT,Zoomcar_EN_CSAT&language=EN&rollBack=&type=EXCEL&schedule='
    }

    activity_report_url = f'https://centrogs.sencall-technologies.com:8443/aheevaManagerServer/api/getAgentActivityReportDownload?tenantDbid={account_id[account]}&fromDate={start_year}-{start_month}-{str(start_day - 1).zfill(2)}T{time_start}:00:00.000Z&toDate={end_year}-{end_month}-{str(end_day + 1).zfill(2)}T{time_end}:59:59.000Z&offset={offset}&tenantName={account}&agentsIDList={agent_ids}&agentGroupDbid=&agentGroupName=AllAgentGroups&language=EN&rollBack=&type=EXCEL'
    # print(activity_report_url)
    inbound_report_url = inbound_urls[account]
    # print(inbound_report_url)
    with requests.Session() as s:
        s.post("https://centrogs.sencall-technologies.com:8443/aheevaManagerServer/api/tenantLogin", login_data)
        home_page = s.get(
            f"https://centrogs.sencall-technologies.com:8443/?{account}#/tenant/{account_id[account]}/home")
        # print(f'https://centrogs.sencall-technologies.com:8443/?{account}#/tenant/{account_id[account]}/home')
        home_page.raise_for_status()
        try:
            print('getting outbound file')
            activity_response = s.get(activity_report_url)
            activity_response.raise_for_status()
            print('getting inbound file')
            inbound_response = s.get(inbound_report_url)
            inbound_response.raise_for_status()
        except requests.HTTPError:
            print(f'\n\nCredentials for {account} corrupted, please check your credentials file.')
            return 1

    print('saving data')
    outbound_file_name = f'{account}_outbound.xlsx'
    inbound_file_name = f'{account}_inbound.xlsx'

    with open(outbound_file_name, 'wb') as outbound_file:
        outbound_file.write(activity_response.content)

    with open(inbound_file_name, 'wb') as inbound_file:
        inbound_file.write(inbound_response.content)

    print('converting to csv')
    # Converting file to csv
    book = openpyxl.load_workbook(outbound_file_name)
    sheet = book['Agent Activity Report']
    sheet.delete_rows(1, 7)
    removeFormatting(sheet)
    book.save('Agent_Activity_Report_Adjusted.xlsx')

    df = pd.read_excel('Agent_Activity_Report_Adjusted.xlsx')
    df.to_csv(f'{account}_outbound.csv', index=None)

    # Converting file to csv
    book = openpyxl.load_workbook(inbound_file_name)
    sheet = book['Inbound Call Life Report']
    sheet.delete_rows(1, 7)
    removeFormatting(sheet)
    book.save('Inbound_Call_Life_Report_Adjusted.xlsx')

    df = pd.read_excel('Inbound_Call_Life_Report_Adjusted.xlsx')
    # print(df)
    df.to_csv(f'{account}_inbound.csv', index=None)

    os.remove(inbound_file_name)
    os.remove(outbound_file_name)
    os.remove('Inbound_Call_Life_Report_Adjusted.xlsx')
    os.remove('Agent_Activity_Report_Adjusted.xlsx')
    print('date updated...')

    print('adding to cache')
    today = datetime.today().date()
    if all([start_date != today, end_date != today]):
        copy(f'{account}_inbound.csv', f'data_cache/{account}_{start_date}_{end_date}_inbound.csv')
        copy(f'{account}_outbound.csv', f'data_cache/{account}_{start_date}_{end_date}_outbound.csv')


if __name__ == '__main__':
    today = datetime.today().date()
    update_data('swvl', today, today)
